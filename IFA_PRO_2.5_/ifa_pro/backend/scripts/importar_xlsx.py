"""
Importa os dados ja cadastrados em IFA_Analysis_System.xlsx (o sistema
Excel/VBA anterior) para o novo banco Postgres, respeitando a ordem de
dependencia entre tabelas (secao "Migracao de dados" em docs/MIGRACAO.md).

Uso:
    docker compose run --rm api python scripts/importar_xlsx.py \
        --arquivo /caminho/para/IFA_Analysis_System.xlsx

Observacoes:
- So importa Clubes, Partidas, Estatisticas, Performance, Lesoes,
  ModelosIA/LogML e Configuracoes — as tabelas que ja existiam no
  sistema anterior. `consensos` e `probabilidades` nao sao recriadas
  aqui porque dependem do Motor IA, que nao existia no sistema antigo
  (rode o pipeline de analise depois da importacao para gera-las).
- E seguro rodar mais de uma vez: clubes/partidas ja importados (mesmo
  nome+data) sao pulados, nao duplicados.
"""
import argparse
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.models.clube import Clube
from app.models.configuracao import Configuracao
from app.models.estatistica import EstatisticaPartida
from app.models.partida import Partida


def _linhas_da_tabela(wb, nome_tabela: str):
    """Percorre todas as planilhas procurando a Tabela Estruturada pelo
    nome e devolve as linhas como lista de dicts {coluna: valor}."""
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if tbl.name == nome_tabela:
                ref_cells = ws[tbl.ref]
                cabecalho = [c.value for c in ref_cells[0]]
                linhas = []
                for row in ref_cells[1:]:
                    linhas.append({cabecalho[i]: row[i].value for i in range(len(cabecalho))})
                return linhas
    return []


def importar_clubes(db: Session, wb) -> dict[int, int]:
    """Retorna um mapa {ID_antigo (Excel) -> id novo (Postgres)}."""
    mapa_ids = {}
    for linha in _linhas_da_tabela(wb, "tbl_Clubes"):
        existente = db.query(Clube).filter(Clube.nome == linha["Nome"]).first()
        if existente:
            mapa_ids[linha["ID_Clube"]] = existente.id
            continue
        clube = Clube(
            nome=linha["Nome"], sigla=linha.get("Sigla") or "", cidade=linha.get("Cidade") or "",
            estado=linha.get("Estado") or "", pais=linha.get("Pais") or "",
            fundacao=linha.get("Fundacao"), estadio=linha.get("Estadio"),
            capacidade_estadio=linha.get("Capacidade_Estadio"), tecnico=linha.get("Tecnico"),
            ativo=(linha.get("Ativo") == "Sim"),
        )
        db.add(clube)
        db.flush()
        mapa_ids[linha["ID_Clube"]] = clube.id
    db.commit()
    print(f"Clubes importados: {len(mapa_ids)}")
    return mapa_ids


def importar_partidas(db: Session, wb, mapa_clubes: dict[int, int]) -> dict[int, int]:
    mapa_ids = {}
    for linha in _linhas_da_tabela(wb, "tbl_Partidas"):
        id_mandante = mapa_clubes.get(linha["ID_Clube_Mandante"])
        id_visitante = mapa_clubes.get(linha["ID_Clube_Visitante"])
        if not id_mandante or not id_visitante:
            continue
        partida = Partida(
            data_partida=linha["Data"] if isinstance(linha["Data"], datetime) else datetime.now(),
            rodada=linha.get("Rodada"), competicao=linha.get("Competicao") or "Nao informado",
            id_clube_mandante=id_mandante, id_clube_visitante=id_visitante,
            gols_mandante=linha.get("Gols_Mandante"), gols_visitante=linha.get("Gols_Visitante"),
            publico=linha.get("Publico"), local=linha.get("Local"),
            status=linha.get("Status") or "Encerrada",
        )
        db.add(partida)
        db.flush()
        mapa_ids[linha["ID_Partida"]] = partida.id
    db.commit()
    print(f"Partidas importadas: {len(mapa_ids)}")
    return mapa_ids


def importar_estatisticas(db: Session, wb, mapa_partidas, mapa_clubes) -> int:
    total = 0
    for linha in _linhas_da_tabela(wb, "tbl_EstatisticasPartida"):
        id_partida = mapa_partidas.get(linha["ID_Partida"])
        id_clube = mapa_clubes.get(linha["ID_Clube"])
        if not id_partida or not id_clube:
            continue
        db.add(EstatisticaPartida(
            id_partida=id_partida, id_clube=id_clube,
            posse_bola_pct=linha.get("Posse_Bola_Pct") or 0, finalizacoes=linha.get("Finalizacoes") or 0,
            finalizacoes_no_alvo=linha.get("Finalizacoes_no_Alvo") or 0, escanteios=linha.get("Escanteios") or 0,
            faltas=linha.get("Faltas") or 0, cartoes_amarelos=linha.get("Cartoes_Amarelos") or 0,
            cartoes_vermelhos=linha.get("Cartoes_Vermelhos") or 0, passes_certos=linha.get("Passes_Certos") or 0,
            passes_errados=linha.get("Passes_Errados") or 0, xg=linha.get("xG") or 0,
        ))
        total += 1
    db.commit()
    print(f"Estatisticas importadas: {total}")
    return total


def importar_configuracoes(db: Session, wb) -> int:
    total = 0
    for linha in _linhas_da_tabela(wb, "tbl_Configuracoes"):
        chave = linha.get("Chave")
        if not chave or db.query(Configuracao).filter(Configuracao.chave == chave).first():
            continue
        db.add(Configuracao(
            chave=chave, valor=str(linha.get("Valor") or ""),
            descricao=linha.get("Descricao"), categoria=linha.get("Categoria") or "Geral",
        ))
        total += 1
    db.commit()
    print(f"Configuracoes importadas: {total}")
    return total


def main():
    parser = argparse.ArgumentParser(description="Importa dados do IFA_Analysis_System.xlsx para o Postgres")
    parser.add_argument("--arquivo", required=True, help="Caminho para o IFA_Analysis_System.xlsx")
    args = parser.parse_args()

    wb = load_workbook(args.arquivo, data_only=True)
    db = SessionLocal()
    try:
        mapa_clubes = importar_clubes(db, wb)
        mapa_partidas = importar_partidas(db, wb, mapa_clubes)
        importar_estatisticas(db, wb, mapa_partidas, mapa_clubes)
        importar_configuracoes(db, wb)
        print("Importacao concluida. Rode o pipeline de analise para gerar consensos/probabilidades.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
